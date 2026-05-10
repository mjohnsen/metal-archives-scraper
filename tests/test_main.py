"""
Tests for main.py — AdaptiveThrottle, _apply_artist_match, and _process_artist.

main.py has module-level side effects (mkdir, logging.basicConfig) that run on
import; these are harmless in the test environment (logs/ already exists,
basicConfig is a no-op if handlers are already configured).
"""
import pytest
from unittest.mock import MagicMock, call, patch
from openpyxl import Workbook

# Import the functions under test directly
from main import AdaptiveThrottle, _apply_artist_match, _fetch_artist_pages, _pick_best_by_type, _process_artist

import metal_archives_scraper.spreadsheet as spreadsheet_module
import metal_archives_scraper.scraper as scraper_module


# ---------------------------------------------------------------------------
# Helpers / shared fixtures
# ---------------------------------------------------------------------------

def _collection_ws():
    """Minimal in-memory Collection worksheet with two unsearched Pharaoh rows."""
    from metal_archives_scraper.spreadsheet import (
        C_ARTIST, C_RELEASE, C_YEAR,
        C_MA_ARTIST_URL, C_MA_RELEASE_URL, C_TYPE,
        C_SEARCHED, C_FOUND, C_REVIEW_FLAG,
        ensure_collection_sheet, ensure_artists_sheet,
        ensure_review_sheet, ensure_not_found_sheet,
    )
    wb = Workbook()
    ws = wb.active
    headers = [
        C_ARTIST, C_RELEASE, C_YEAR,
        C_MA_ARTIST_URL, C_MA_RELEASE_URL, C_TYPE,
        C_SEARCHED, C_FOUND, C_REVIEW_FLAG,
    ]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    ws.cell(row=2, column=1, value="Pharaoh")
    ws.cell(row=2, column=2, value="After the Fire")
    ws.cell(row=3, column=1, value="Pharaoh")
    ws.cell(row=3, column=2, value="The Longest Night")
    ws.title = "Collection"
    ws_artists = ensure_artists_sheet(wb)
    ws_review = ensure_review_sheet(wb)
    ws_not_found = ensure_not_found_sheet(wb)
    return wb, ws, ws_artists, ws_review, ws_not_found


def _pharaoh_ranked(artist_dict, collection_titles):
    """A ranked entry as returned by rank_artists_by_releases."""
    matched = [
        {
            "collection_title": t,
            "discography_title": t,
            "discography_url": f"https://ma.com/albums/{t.replace(' ', '_')}/1",
            "discography_type": "Full-length",
            "discography_year": "2008",
            "match_score": 100,
        }
        for t in collection_titles
    ]
    return {
        "artist": artist_dict,
        "matched_releases": matched,
        "match_count": len(matched),
        "avg_score": 100.0,
    }


def _make_throttle():
    """A zero-interval AdaptiveThrottle that never sleeps."""
    return AdaptiveThrottle(initial=0.0, min_interval=0.0)


@pytest.fixture
def pharaoh_artist_dict():
    return {
        "name": "Pharaoh",
        "url": "https://www.metal-archives.com/bands/Pharaoh/2801",
        "band_id": "2801",
        "country": "United States",
        "location": "Philadelphia, PA",
        "status": "Active",
        "formed_in": "2000",
        "genre": "Power/Progressive Metal",
        "themes": "Philosophy",
        "label": "Cruz Del Sur Music",
        "years_active": "2000-present",
        "discography": [
            {
                "title": "After the Fire",
                "url": "https://ma.com/albums/Pharaoh/After_the_Fire/123",
                "type": "Full-length",
                "year": "2008",
            },
            {
                "title": "The Longest Night",
                "url": "https://ma.com/albums/Pharaoh/The_Longest_Night/456",
                "type": "Full-length",
                "year": "2006",
            },
        ],
    }


# ---------------------------------------------------------------------------
# AdaptiveThrottle
# ---------------------------------------------------------------------------

class TestAdaptiveThrottle:
    def test_sleeps_when_interval_not_elapsed(self, monkeypatch):
        mock_sleep = MagicMock()
        monkeypatch.setattr("main.time.sleep", mock_sleep)
        monkeypatch.setattr("main.time.time", MagicMock(return_value=0.0))
        monkeypatch.setattr("main.random.uniform", MagicMock(return_value=0.0))

        throttle = AdaptiveThrottle(initial=30.0, min_interval=5.0)
        throttle._last_time = 0.0
        throttle.wait("test")

        mock_sleep.assert_called_once()
        assert mock_sleep.call_args[0][0] == pytest.approx(30.0, abs=0.01)

    def test_does_not_sleep_when_interval_already_elapsed(self, monkeypatch):
        mock_sleep = MagicMock()
        monkeypatch.setattr("main.time.sleep", mock_sleep)
        monkeypatch.setattr("main.time.time", MagicMock(return_value=100.0))
        monkeypatch.setattr("main.random.uniform", MagicMock(return_value=0.0))

        throttle = AdaptiveThrottle(initial=30.0, min_interval=5.0)
        throttle._last_time = 0.0  # 100s elapsed > 30s interval
        throttle.wait("test")

        mock_sleep.assert_not_called()

    def test_success_decreases_interval_after_threshold(self, monkeypatch):
        monkeypatch.setattr("main.time.time", MagicMock(return_value=0.0))

        throttle = AdaptiveThrottle(initial=30.0, min_interval=5.0,
                                    decrease_factor=0.5, successes_to_decrease=3)
        throttle.success()
        throttle.success()
        assert throttle.interval == pytest.approx(30.0)  # not yet
        throttle.success()
        assert throttle.interval == pytest.approx(15.0)  # 30 * 0.5

    def test_success_does_not_go_below_minimum(self, monkeypatch):
        monkeypatch.setattr("main.time.time", MagicMock(return_value=0.0))

        throttle = AdaptiveThrottle(initial=6.0, min_interval=5.0,
                                    decrease_factor=0.5, successes_to_decrease=1)
        throttle.success()
        assert throttle.interval == pytest.approx(5.0)  # clamped to min

    def test_reset_restores_baseline(self, monkeypatch):
        monkeypatch.setattr("main.time.time", MagicMock(return_value=0.0))

        throttle = AdaptiveThrottle(initial=30.0, min_interval=5.0,
                                    decrease_factor=0.5, successes_to_decrease=1)
        throttle.success()  # drops to 15.0
        throttle.reset()
        assert throttle.interval == pytest.approx(30.0)

    def test_reset_clears_consecutive_successes(self, monkeypatch):
        monkeypatch.setattr("main.time.time", MagicMock(return_value=0.0))

        throttle = AdaptiveThrottle(initial=30.0, min_interval=5.0,
                                    decrease_factor=0.5, successes_to_decrease=3)
        throttle.success()
        throttle.success()
        throttle.reset()
        # After reset, need another full threshold to trigger a decrease
        throttle.success()
        throttle.success()
        assert throttle.interval == pytest.approx(30.0)  # not yet decreased

    def test_jitter_adds_variance_to_wait(self, monkeypatch):
        sleep_calls = []
        monkeypatch.setattr("main.time.sleep", lambda s: sleep_calls.append(s))
        monkeypatch.setattr("main.time.time", MagicMock(return_value=0.0))

        # uniform returns +1 then -1
        uniform_values = iter([1.0, -1.0])
        monkeypatch.setattr("main.random.uniform", lambda a, b: next(uniform_values))

        throttle = AdaptiveThrottle(initial=20.0, min_interval=5.0, jitter=0.25)
        throttle._last_time = 0.0
        throttle.wait("test")  # jitter_offset = 20*0.25*1 = +5 → waits 25s
        throttle.wait("test")  # jitter_offset = 20*0.25*(-1) = -5 → waits 15s

        assert sleep_calls[0] == pytest.approx(25.0, abs=0.01)
        assert sleep_calls[1] == pytest.approx(15.0, abs=0.01)


# ---------------------------------------------------------------------------
# _apply_artist_match
# ---------------------------------------------------------------------------

class TestApplyArtistMatch:
    def test_marks_matched_releases_as_found(self, pharaoh_artist_dict):
        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        ranked = _pharaoh_ranked(pharaoh_artist_dict, ["After the Fire"])

        _apply_artist_match(ranked, ["After the Fire", "The Longest Night"],
                            "Pharaoh", ws, ws_artists, ws_review, ws_not_found)

        from metal_archives_scraper.spreadsheet import C_FOUND, C_SEARCHED, _get_col_map
        col_map = _get_col_map(ws)
        row2_searched = ws.cell(row=2, column=col_map[C_SEARCHED]).value
        row2_found = ws.cell(row=2, column=col_map[C_FOUND]).value
        assert row2_searched is True
        assert row2_found is True

    def test_marks_unmatched_releases_as_not_found(self, pharaoh_artist_dict):
        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        ranked = _pharaoh_ranked(pharaoh_artist_dict, ["After the Fire"])

        _apply_artist_match(ranked, ["After the Fire", "The Longest Night"],
                            "Pharaoh", ws, ws_artists, ws_review, ws_not_found)

        from metal_archives_scraper.spreadsheet import C_FOUND, C_SEARCHED, _get_col_map
        col_map = _get_col_map(ws)
        row3_found = ws.cell(row=3, column=col_map[C_FOUND]).value
        assert row3_found is False

    def test_adds_unmatched_releases_to_not_found_sheet(self, pharaoh_artist_dict):
        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        ranked = _pharaoh_ranked(pharaoh_artist_dict, ["After the Fire"])

        _apply_artist_match(ranked, ["After the Fire", "The Longest Night"],
                            "Pharaoh", ws, ws_artists, ws_review, ws_not_found)

        not_found_releases = [
            ws_not_found.cell(row=r, column=2).value
            for r in range(2, ws_not_found.max_row + 1)
        ]
        assert "The Longest Night" in not_found_releases

    def test_calls_update_artist_row(self, pharaoh_artist_dict, monkeypatch):
        mock_update = MagicMock()
        monkeypatch.setattr(spreadsheet_module, "update_artist_row", mock_update)
        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        ranked = _pharaoh_ranked(pharaoh_artist_dict, ["After the Fire"])

        _apply_artist_match(ranked, ["After the Fire"], "Pharaoh",
                            ws, ws_artists, ws_review, ws_not_found)

        mock_update.assert_called_once_with(ws_artists, pharaoh_artist_dict)

    def test_flags_duplicate_titled_releases_for_review(self, pharaoh_artist_dict):
        """Duplicate-titled discography entries: sets needs_review and adds review sheet entry."""
        pharaoh_artist_dict["discography"].append({
            "title": "After the Fire",
            "url": "https://ma.com/albums/Pharaoh/After_the_Fire/999",
            "type": "Single",
            "year": "2007",
        })
        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        ranked = _pharaoh_ranked(pharaoh_artist_dict, ["After the Fire"])

        _apply_artist_match(ranked, ["After the Fire"], "Pharaoh",
                            ws, ws_artists, ws_review, ws_not_found)

        from metal_archives_scraper.spreadsheet import C_REVIEW_FLAG, _get_col_map
        col_map = _get_col_map(ws)
        assert ws.cell(row=2, column=col_map[C_REVIEW_FLAG]).value is True
        assert ws_review.max_row >= 2

    def test_picks_higher_precedence_type_when_duplicates_exist(self, pharaoh_artist_dict):
        """Full-length takes precedence over Single when both share a title."""
        pharaoh_artist_dict["discography"].append({
            "title": "After the Fire",
            "url": "https://ma.com/albums/Pharaoh/After_the_Fire/999",
            "type": "Single",
            "year": "2007",
        })
        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        ranked = _pharaoh_ranked(pharaoh_artist_dict, ["After the Fire"])

        _apply_artist_match(ranked, ["After the Fire"], "Pharaoh",
                            ws, ws_artists, ws_review, ws_not_found)

        from metal_archives_scraper.spreadsheet import C_TYPE, _get_col_map
        col_map = _get_col_map(ws)
        assert ws.cell(row=2, column=col_map[C_TYPE]).value == "Full-length"

    def test_falls_back_to_first_when_same_type_duplicates(self, pharaoh_artist_dict):
        """When duplicate titles share the same type, the first entry in the discography is used."""
        pharaoh_artist_dict["discography"].append({
            "title": "After the Fire",
            "url": "https://ma.com/albums/Pharaoh/After_the_Fire/999",
            "type": "Full-length",
            "year": "2009",
        })
        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        ranked = _pharaoh_ranked(pharaoh_artist_dict, ["After the Fire"])

        _apply_artist_match(ranked, ["After the Fire"], "Pharaoh",
                            ws, ws_artists, ws_review, ws_not_found)

        from metal_archives_scraper.spreadsheet import C_MA_RELEASE_URL, _get_col_map
        col_map = _get_col_map(ws)
        # First entry in discography is used (url ending /123, not /999)
        assert ws.cell(row=2, column=col_map[C_MA_RELEASE_URL]).value == \
            "https://ma.com/albums/Pharaoh/After_the_Fire/123"


# ---------------------------------------------------------------------------
# _pick_best_by_type
# ---------------------------------------------------------------------------

class TestPickBestByType:
    def _entry(self, title, release_type, url_suffix="1"):
        return {"title": title, "url": f"https://ma.com/{url_suffix}", "type": release_type, "year": "2000"}

    def test_returns_single_entry_unchanged(self):
        entry = self._entry("Album", "EP")
        assert _pick_best_by_type([entry]) is entry

    def test_full_length_beats_ep(self):
        ep = self._entry("Album", "EP", "2")
        fl = self._entry("Album", "Full-length", "1")
        assert _pick_best_by_type([ep, fl])["type"] == "Full-length"

    def test_ep_beats_single(self):
        single = self._entry("Album", "Single", "2")
        ep = self._entry("Album", "EP", "1")
        assert _pick_best_by_type([single, ep])["type"] == "EP"

    def test_single_beats_live_album(self):
        live = self._entry("Album", "Live album", "2")
        single = self._entry("Album", "Single", "1")
        assert _pick_best_by_type([single, live])["type"] == "Single"

    def test_unknown_type_loses_to_any_known_type(self):
        unknown = self._entry("Album", "Video", "2")
        demo = self._entry("Album", "Demo", "1")
        assert _pick_best_by_type([unknown, demo])["type"] == "Demo"

    def test_same_type_returns_first_in_list(self):
        first = self._entry("Album", "Full-length", "1")
        second = self._entry("Album", "Full-length", "2")
        assert _pick_best_by_type([first, second]) is first


# ---------------------------------------------------------------------------
# _fetch_artist_pages
# ---------------------------------------------------------------------------

class TestFetchArtistPages:
    def test_returns_artist_dicts_for_each_url(self, pharaoh_artist_dict, monkeypatch):
        monkeypatch.setattr(scraper_module, "get_artist_info",
                            MagicMock(return_value=pharaoh_artist_dict))
        results = {"https://ma.com/bands/Pharaoh/1": {}, "https://ma.com/bands/Pharaoh/2": {}}
        artist_dicts, fetch_errors = _fetch_artist_pages(results, _make_throttle())
        assert len(artist_dicts) == 2
        assert fetch_errors == 0

    def test_counts_non_runtime_errors_as_fetch_errors(self, monkeypatch):
        monkeypatch.setattr(scraper_module, "get_artist_info",
                            MagicMock(side_effect=Exception("Network error")))
        artist_dicts, fetch_errors = _fetch_artist_pages({"https://ma.com/1": {}}, _make_throttle())
        assert artist_dicts == []
        assert fetch_errors == 1

    def test_propagates_runtime_error(self, monkeypatch):
        monkeypatch.setattr(scraper_module, "get_artist_info",
                            MagicMock(side_effect=RuntimeError("Cloudflare timeout")))
        with pytest.raises(RuntimeError):
            _fetch_artist_pages({"https://ma.com/1": {}}, _make_throttle())

    def test_calls_throttle_wait_and_success_per_url(self, pharaoh_artist_dict, monkeypatch):
        monkeypatch.setattr(scraper_module, "get_artist_info",
                            MagicMock(return_value=pharaoh_artist_dict))
        throttle = MagicMock(spec=AdaptiveThrottle)
        results = {"https://ma.com/1": {}, "https://ma.com/2": {}}
        _fetch_artist_pages(results, throttle)
        assert throttle.wait.call_count == 2
        assert throttle.success.call_count == 2

    def test_does_not_call_success_on_fetch_error(self, monkeypatch):
        monkeypatch.setattr(scraper_module, "get_artist_info",
                            MagicMock(side_effect=Exception("err")))
        throttle = MagicMock(spec=AdaptiveThrottle)
        _fetch_artist_pages({"https://ma.com/1": {}}, throttle)
        throttle.success.assert_not_called()


# ---------------------------------------------------------------------------
# _process_artist
# ---------------------------------------------------------------------------

class TestProcessArtist:
    """Integration-level tests for the main artist-processing loop.

    All network I/O (scraper.*), spreadsheet writes, and throttling are mocked.
    """

    @pytest.fixture(autouse=True)
    def mock_save(self, monkeypatch):
        monkeypatch.setattr(spreadsheet_module, "save_workbook", MagicMock())

    def _run(self, monkeypatch, search_results, artist_info=None, pharaoh_dict=None):
        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()

        if artist_info is None:
            artist_info = pharaoh_dict or {}

        monkeypatch.setattr(scraper_module, "search_artist_release",
                            MagicMock(return_value=search_results))
        monkeypatch.setattr(scraper_module, "get_artist_info",
                            MagicMock(return_value=artist_info))

        _process_artist(
            artist_name="Pharaoh",
            release_titles=["After the Fire", "The Longest Night"],
            ws_collection=ws,
            ws_artists=ws_artists,
            ws_review=ws_review,
            ws_not_found=ws_not_found,
            wb=wb,
            path="/fake/path.xlsx",
            query_throttle=_make_throttle(),
            artist_throttle=_make_throttle(),
        )
        return ws, ws_not_found

    def test_propagates_runtime_error_from_get_artist_info(self, monkeypatch,
                                                             pharaoh_artist_dict):
        """A RuntimeError from get_artist_info (e.g. Cloudflare timeout) must propagate."""
        monkeypatch.setattr(scraper_module, "search_artist_release",
                            MagicMock(return_value={
                                pharaoh_artist_dict["url"]: {}
                            }))
        monkeypatch.setattr(scraper_module, "get_artist_info",
                            MagicMock(side_effect=RuntimeError("Timed out waiting for Cloudflare")))

        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        with pytest.raises(RuntimeError, match="Cloudflare"):
            _process_artist(
                artist_name="Pharaoh",
                release_titles=["After the Fire"],
                ws_collection=ws, ws_artists=ws_artists, ws_review=ws_review,
                ws_not_found=ws_not_found, wb=wb, path="/fake/path.xlsx",
                query_throttle=_make_throttle(),
                artist_throttle=_make_throttle(),
            )

    def test_st_title_is_expanded_to_artist_name_before_search(self, monkeypatch,
                                                                pharaoh_artist_dict):
        """An 's/t' release title should be substituted with the artist name before querying."""
        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        # Override row 2 to have an s/t title
        ws.cell(row=2, column=2, value="s/t")

        search_mock = MagicMock(return_value={})
        monkeypatch.setattr(scraper_module, "search_artist_release", search_mock)

        _process_artist(
            artist_name="Pharaoh",
            release_titles=["s/t", "The Longest Night"],
            ws_collection=ws, ws_artists=ws_artists, ws_review=ws_review,
            ws_not_found=ws_not_found, wb=wb, path="/fake/path.xlsx",
            query_throttle=_make_throttle(),
            artist_throttle=_make_throttle(),
        )

        # The cell should now hold the artist name, not "s/t"
        assert ws.cell(row=2, column=2).value == "Pharaoh"

        # search_artist_release should have been called with the artist name, not "s/t"
        searched_titles = [c.args[1] for c in search_mock.call_args_list]
        assert "Pharaoh" in searched_titles
        assert "s/t" not in searched_titles

        # A review entry should record the substitution
        assert ws_review.max_row >= 2
        review_issues = [ws_review.cell(row=r, column=3).value
                         for r in range(2, ws_review.max_row + 1)]
        assert any("s/t changed to artist name" in (v or "") for v in review_issues)

    def test_marks_release_not_found_when_no_search_results(self, monkeypatch,
                                                              pharaoh_artist_dict):
        ws, ws_not_found = self._run(
            monkeypatch, search_results={}, pharaoh_dict=pharaoh_artist_dict
        )
        from metal_archives_scraper.spreadsheet import C_FOUND, C_SEARCHED, _get_col_map
        col_map = _get_col_map(ws)
        # Both releases should be marked searched=True, found=False eventually
        found_vals = [
            ws.cell(row=r, column=col_map[C_FOUND]).value
            for r in range(2, ws.max_row + 1)
        ]
        assert all(v is False for v in found_vals if v is not None)

    def test_fetches_artist_page_for_each_unique_url(self, monkeypatch, pharaoh_artist_dict):
        mock_get_info = MagicMock(return_value=pharaoh_artist_dict)
        monkeypatch.setattr(scraper_module, "get_artist_info", mock_get_info)
        monkeypatch.setattr(scraper_module, "search_artist_release",
                            MagicMock(return_value={
                                pharaoh_artist_dict["url"]: {"artist_url": pharaoh_artist_dict["url"]}
                            }))
        monkeypatch.setattr(spreadsheet_module, "save_workbook", MagicMock())

        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        _process_artist(
            artist_name="Pharaoh",
            release_titles=["After the Fire"],
            ws_collection=ws, ws_artists=ws_artists, ws_review=ws_review,
            ws_not_found=ws_not_found, wb=wb, path="/fake/path.xlsx",
            query_throttle=_make_throttle(),
            artist_throttle=_make_throttle(),
        )
        mock_get_info.assert_called()

    def test_falls_back_to_fuzzy_search_when_exact_returns_empty(self, monkeypatch,
                                                                   pharaoh_artist_dict):
        search_mock = MagicMock(side_effect=[{}, {}])  # both exact and fuzzy return empty
        monkeypatch.setattr(scraper_module, "search_artist_release", search_mock)
        monkeypatch.setattr(spreadsheet_module, "save_workbook", MagicMock())

        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        _process_artist(
            artist_name="Pharaoh",
            release_titles=["After the Fire"],
            ws_collection=ws, ws_artists=ws_artists, ws_review=ws_review,
            ws_not_found=ws_not_found, wb=wb, path="/fake/path.xlsx",
            query_throttle=_make_throttle(),
            artist_throttle=_make_throttle(),
        )
        # Should have been called twice: exact=True then exact=False
        calls = search_mock.call_args_list
        exact_flags = [c.kwargs.get("exact_band", c.args[2] if len(c.args) > 2 else True)
                       for c in calls]
        assert True in exact_flags
        assert False in exact_flags

    def test_adds_review_entry_when_multiple_artists_match_same_release(self, monkeypatch,
                                                                          pharaoh_artist_dict):
        other = dict(pharaoh_artist_dict)
        other["url"] = "https://ma.com/bands/Pharaoh/9999"
        other["name"] = "Pharaoh (DE)"

        search_results = {
            pharaoh_artist_dict["url"]: {},
            other["url"]: {},
        }
        monkeypatch.setattr(scraper_module, "search_artist_release",
                            MagicMock(return_value=search_results))

        def fake_get_info(url):
            return pharaoh_artist_dict if url == pharaoh_artist_dict["url"] else other
        monkeypatch.setattr(scraper_module, "get_artist_info", fake_get_info)
        monkeypatch.setattr(spreadsheet_module, "save_workbook", MagicMock())

        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        _process_artist(
            artist_name="Pharaoh",
            release_titles=["After the Fire"],
            ws_collection=ws, ws_artists=ws_artists, ws_review=ws_review,
            ws_not_found=ws_not_found, wb=wb, path="/fake/path.xlsx",
            query_throttle=_make_throttle(),
            artist_throttle=_make_throttle(),
        )
        # Review sheet should have at least one entry for the ambiguous release
        assert ws_review.max_row >= 2

    def test_adds_review_entry_on_partial_match(self, monkeypatch, pharaoh_artist_dict):
        """When search finds results but fuzzy ranking yields zero matches, add a review entry
        and mark found=False — but do NOT add the release to the Not Found sheet."""
        search_results = {
            pharaoh_artist_dict["url"]: {
                "artist_url": pharaoh_artist_dict["url"],
                "release_url": "https://ma.com/albums/Pharaoh/Realm_of_Chaos/99",
            }
        }
        monkeypatch.setattr(scraper_module, "search_artist_release",
                            MagicMock(return_value=search_results))
        monkeypatch.setattr(scraper_module, "get_artist_info",
                            MagicMock(return_value=pharaoh_artist_dict))
        monkeypatch.setattr(
            scraper_module,
            "rank_artists_by_releases",
            MagicMock(return_value=[{
                "artist": pharaoh_artist_dict,
                "matched_releases": [],
                "match_count": 0,
                "avg_score": 0,
            }]),
        )

        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        _process_artist(
            artist_name="Pharaoh",
            release_titles=["After the Fire", "The Longest Night"],
            ws_collection=ws, ws_artists=ws_artists, ws_review=ws_review,
            ws_not_found=ws_not_found, wb=wb, path="/fake/path.xlsx",
            query_throttle=_make_throttle(),
            artist_throttle=_make_throttle(),
        )

        from metal_archives_scraper.spreadsheet import C_FOUND, C_SEARCHED, _get_col_map
        col_map = _get_col_map(ws)

        # Both releases should appear on the review sheet
        assert ws_review.max_row >= 2
        review_issues = [ws_review.cell(row=r, column=3).value
                         for r in range(2, ws_review.max_row + 1)]
        assert any("Partial match" in (v or "") for v in review_issues)

        # Both releases should be marked found=False
        found_vals = [ws.cell(row=r, column=col_map[C_FOUND]).value
                      for r in range(2, ws.max_row + 1)]
        assert all(v is False for v in found_vals if v is not None)

        # Neither release should land on the Not Found sheet
        assert ws_not_found.max_row == 1

    def test_save_workbook_called_exactly_once(self, monkeypatch, pharaoh_artist_dict):
        """Workbook should be saved once per artist regardless of how many releases are processed."""
        mock_save = MagicMock()
        monkeypatch.setattr(spreadsheet_module, "save_workbook", mock_save)
        monkeypatch.setattr(scraper_module, "search_artist_release",
                            MagicMock(return_value={}))

        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        _process_artist(
            artist_name="Pharaoh",
            release_titles=["After the Fire", "The Longest Night"],
            ws_collection=ws, ws_artists=ws_artists, ws_review=ws_review,
            ws_not_found=ws_not_found, wb=wb, path="/fake/path.xlsx",
            query_throttle=_make_throttle(),
            artist_throttle=_make_throttle(),
        )
        mock_save.assert_called_once()

    def test_returns_none(self, monkeypatch, pharaoh_artist_dict):
        """_process_artist now returns None; timing is tracked inside AdaptiveThrottle."""
        monkeypatch.setattr(scraper_module, "search_artist_release",
                            MagicMock(return_value={}))

        wb, ws, ws_artists, ws_review, ws_not_found = _collection_ws()
        result = _process_artist(
            artist_name="Pharaoh",
            release_titles=["After the Fire"],
            ws_collection=ws, ws_artists=ws_artists, ws_review=ws_review,
            ws_not_found=ws_not_found, wb=wb, path="/fake/path.xlsx",
            query_throttle=_make_throttle(),
            artist_throttle=_make_throttle(),
        )
        assert result is None
