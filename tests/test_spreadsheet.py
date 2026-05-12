import json

import pytest
from openpyxl import Workbook

from metal_archives_scraper.spreadsheet import (
    C_ARTIST,
    C_FOUND,
    C_MA_ARTIST_URL,
    C_MA_RELEASE_URL,
    C_RELEASE,
    C_REVIEW_FLAG,
    C_SEARCHED,
    C_TYPE,
    C_YEAR,
    C_GENRE,
    COL_ARTISTS_ARTIST,
    COL_ARTISTS_DISAMBIG,
    COL_ARTISTS_LOCATION,
    COL_ARTISTS_MA_INFO,
    COL_ARTISTS_MA_URL,
    COL_NOT_FOUND_ARTIST,
    COL_NOT_FOUND_RELEASE,
    COL_REVIEW_ARTIST,
    COL_REVIEW_ISSUE,
    COL_REVIEW_MA_URLS,
    COL_REVIEW_RELEASE,
    COLLECTION_ADDED,
    COLLECTION_ORDERED,
    STATS_SHEET,
    _compute_disambiguations,
    _col_index,
    _ensure_column,
    _get_col_map,
    add_not_found_entry,
    add_review_entry,
    ensure_artists_sheet,
    ensure_collection_sheet,
    ensure_not_found_sheet,
    ensure_review_sheet,
    expand_st_titles,
    find_duplicate_rows,
    get_unsearched_artists,
    open_workbook,
    pick_random_artist,
    save_workbook,
    update_artist_row,
    update_release_row,
    update_stats_sheet,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _wb_with_headers(*headers):
    """Create a workbook whose first sheet has the given column headers."""
    wb = Workbook()
    ws = wb.active
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    return wb


def _add_row(ws, col_map: dict, **kwargs):
    """Append a data row to ws using the column map."""
    row = ws.max_row + 1
    for name, val in kwargs.items():
        col = col_map.get(name)
        if col:
            ws.cell(row=row, column=col, value=val)


def _col_val(ws, row: int, col_name: str) -> object:
    """Read a cell value by column header name."""
    for cell in ws[1]:
        if cell.value == col_name:
            return ws.cell(row=row, column=cell.column).value
    return None


# ---------------------------------------------------------------------------
# _col_index
# ---------------------------------------------------------------------------

class TestColIndex:
    def test_finds_existing_column(self):
        wb = _wb_with_headers("Artist", "Release")
        assert _col_index(wb.active, "Artist") == 1
        assert _col_index(wb.active, "Release") == 2

    def test_returns_none_for_missing_column(self):
        wb = _wb_with_headers("Artist", "Release")
        assert _col_index(wb.active, "Nonexistent") is None


# ---------------------------------------------------------------------------
# _ensure_column
# ---------------------------------------------------------------------------

class TestEnsureColumn:
    def test_returns_existing_column_index(self):
        wb = _wb_with_headers("Artist", "Release")
        idx = _ensure_column(wb.active, "Artist")
        assert idx == 1

    def test_appends_missing_column(self):
        wb = _wb_with_headers("Artist", "Release")
        idx = _ensure_column(wb.active, "New Column")
        assert idx == 3
        assert wb.active.cell(row=1, column=3).value == "New Column"

    def test_does_not_duplicate_existing_column(self):
        wb = _wb_with_headers("Artist")
        _ensure_column(wb.active, "Artist")
        assert wb.active.max_column == 1


# ---------------------------------------------------------------------------
# open_workbook
# ---------------------------------------------------------------------------

class TestOpenWorkbook:
    def test_raises_if_artist_column_missing(self, tmp_path):
        wb = _wb_with_headers("Release", "Year")
        path = str(tmp_path / "test.xlsx")
        wb.save(path)
        with pytest.raises(ValueError, match="Artist"):
            open_workbook(path)

    def test_raises_if_release_column_missing(self, tmp_path):
        wb = _wb_with_headers("Artist", "Year")
        path = str(tmp_path / "test.xlsx")
        wb.save(path)
        with pytest.raises(ValueError, match="Release"):
            open_workbook(path)

    def test_succeeds_with_required_columns(self, tmp_path):
        wb = _wb_with_headers("Artist", "Release")
        path = str(tmp_path / "test.xlsx")
        wb.save(path)
        loaded = open_workbook(path)
        assert loaded is not None

    def test_succeeds_with_extra_columns(self, tmp_path):
        wb = _wb_with_headers("Artist", "Release", "Year", "Genre", "Extra")
        path = str(tmp_path / "test.xlsx")
        wb.save(path)
        loaded = open_workbook(path)
        assert loaded is not None


# ---------------------------------------------------------------------------
# ensure_collection_sheet
# ---------------------------------------------------------------------------

class TestEnsureCollectionSheet:
    def test_renames_first_sheet_to_collection(self):
        wb = _wb_with_headers("Artist", "Release")
        wb.active.title = "Sheet1"
        ensure_collection_sheet(wb)
        assert wb.worksheets[0].title == "Collection"

    def test_does_not_rename_if_already_collection(self):
        wb = _wb_with_headers("Artist", "Release")
        wb.active.title = "Collection"
        ensure_collection_sheet(wb)
        assert wb.active.title == "Collection"

    def test_adds_all_required_new_columns(self):
        wb = _wb_with_headers("Artist", "Release")
        ws = ensure_collection_sheet(wb)
        headers = {cell.value for cell in ws[1]}
        for col_name in COLLECTION_ADDED + COLLECTION_ORDERED:
            assert col_name in headers, f"Missing column: {col_name}"

    def test_inserts_year_genre_type_in_order_after_release(self):
        wb = _wb_with_headers("Artist", "Release")
        ws = ensure_collection_sheet(wb)
        headers = [cell.value for cell in ws[1] if cell.value]
        release_idx = headers.index(C_RELEASE)
        year_idx = headers.index(C_YEAR)
        genre_idx = headers.index(C_GENRE)
        type_idx = headers.index(C_TYPE)
        assert release_idx < year_idx < genre_idx < type_idx

    def test_inserts_missing_year_before_existing_genre(self):
        wb = _wb_with_headers("Artist", "Release", C_GENRE)
        ws = ensure_collection_sheet(wb)
        headers = [cell.value for cell in ws[1] if cell.value]
        assert headers.index(C_YEAR) < headers.index(C_GENRE)

    def test_inserts_missing_genre_between_existing_year_and_type(self):
        wb = _wb_with_headers("Artist", "Release", C_YEAR, C_TYPE)
        ws = ensure_collection_sheet(wb)
        headers = [cell.value for cell in ws[1] if cell.value]
        assert headers.index(C_YEAR) < headers.index(C_GENRE) < headers.index(C_TYPE)

    def test_data_rows_preserved_after_column_insertion(self):
        wb = _wb_with_headers("Artist", "Release")
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        ws = ensure_collection_sheet(wb)
        col_map = _get_col_map(ws)
        assert ws.cell(row=2, column=col_map[C_ARTIST]).value == "Pharaoh"
        assert ws.cell(row=2, column=col_map[C_RELEASE]).value == "After the Fire"

    def test_does_not_modify_existing_columns(self):
        wb = _wb_with_headers("Artist", "Release", "Year", "Genre")
        ws = ensure_collection_sheet(wb)
        # Existing columns stay in their original positions
        assert ws.cell(row=1, column=1).value == "Artist"
        assert ws.cell(row=1, column=2).value == "Release"
        assert ws.cell(row=1, column=3).value == "Year"
        assert ws.cell(row=1, column=4).value == "Genre"

    def test_does_not_add_column_that_already_exists(self):
        # Pre-add one of the COLLECTION_ADDED columns
        wb = _wb_with_headers("Artist", "Release", C_SEARCHED)
        ws = ensure_collection_sheet(wb)
        # C_SEARCHED should appear exactly once
        searched_count = sum(1 for cell in ws[1] if cell.value == C_SEARCHED)
        assert searched_count == 1

    def test_does_not_duplicate_ordered_columns_already_present(self):
        wb = _wb_with_headers("Artist", "Release", C_YEAR, C_GENRE, C_TYPE)
        ws = ensure_collection_sheet(wb)
        headers = [cell.value for cell in ws[1] if cell.value]
        for col_name in COLLECTION_ORDERED:
            assert headers.count(col_name) == 1, f"Duplicate column: {col_name}"


# ---------------------------------------------------------------------------
# ensure_artists_sheet
# ---------------------------------------------------------------------------

class TestEnsureArtistsSheet:
    def test_creates_artists_sheet_when_absent(self):
        wb = _wb_with_headers("Artist", "Release")
        ensure_artists_sheet(wb)
        assert "Artists" in wb.sheetnames

    def test_artists_sheet_has_correct_headers(self):
        wb = _wb_with_headers("Artist", "Release")
        ws = ensure_artists_sheet(wb)
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for expected in [COL_ARTISTS_ARTIST, COL_ARTISTS_DISAMBIG,
                         COL_ARTISTS_LOCATION, COL_ARTISTS_MA_URL, COL_ARTISTS_MA_INFO]:
            assert expected in headers

    def test_returns_existing_sheet_when_already_present(self):
        wb = _wb_with_headers("Artist", "Release")
        wb.create_sheet("Artists")
        wb["Artists"].cell(row=1, column=1, value="My Custom Header")
        ws = ensure_artists_sheet(wb)
        # Should return the existing sheet (not create a new one)
        assert ws.title == "Artists"
        assert wb.sheetnames.count("Artists") == 1


# ---------------------------------------------------------------------------
# ensure_review_sheet
# ---------------------------------------------------------------------------

class TestEnsureReviewSheet:
    def test_creates_review_sheet_when_absent(self):
        wb = _wb_with_headers("Artist", "Release")
        ensure_review_sheet(wb)
        assert "Review" in wb.sheetnames

    def test_review_sheet_has_correct_headers(self):
        wb = _wb_with_headers("Artist", "Release")
        ws = ensure_review_sheet(wb)
        headers = {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}
        for h in [COL_REVIEW_ARTIST, COL_REVIEW_RELEASE, COL_REVIEW_ISSUE, COL_REVIEW_MA_URLS]:
            assert h in headers

    def test_returns_existing_sheet(self):
        wb = _wb_with_headers("Artist", "Release")
        wb.create_sheet("Review")
        ws = ensure_review_sheet(wb)
        assert wb.sheetnames.count("Review") == 1


# ---------------------------------------------------------------------------
# ensure_not_found_sheet
# ---------------------------------------------------------------------------

class TestEnsureNotFoundSheet:
    def test_creates_not_found_sheet_when_absent(self):
        wb = _wb_with_headers("Artist", "Release")
        ensure_not_found_sheet(wb)
        assert "Not Found" in wb.sheetnames

    def test_not_found_sheet_has_correct_headers(self):
        wb = _wb_with_headers("Artist", "Release")
        ws = ensure_not_found_sheet(wb)
        headers = {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}
        assert COL_NOT_FOUND_ARTIST in headers
        assert COL_NOT_FOUND_RELEASE in headers


# ---------------------------------------------------------------------------
# get_unsearched_artists
# ---------------------------------------------------------------------------

class TestGetUnsearchedArtists:
    def _make_ws(self):
        """Collection sheet with searched and unsearched rows."""
        wb = _wb_with_headers(
            C_ARTIST, C_RELEASE, C_YEAR, C_SEARCHED, C_FOUND
        )
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")

        ws.cell(row=3, column=1, value="Pharaoh")
        ws.cell(row=3, column=2, value="The Longest Night")

        ws.cell(row=4, column=1, value="Iced Earth")
        ws.cell(row=4, column=2, value="Framing Armageddon")
        ws.cell(row=4, column=4, value=True)  # Searched = True
        return ws

    def test_returns_unsearched_artists(self):
        ws = self._make_ws()
        result = get_unsearched_artists(ws)
        assert "Pharaoh" in result

    def test_excludes_already_searched_rows(self):
        ws = self._make_ws()
        result = get_unsearched_artists(ws)
        assert "Iced Earth" not in result

    def test_groups_multiple_releases_under_same_artist(self):
        ws = self._make_ws()
        result = get_unsearched_artists(ws)
        assert "After the Fire" in result["Pharaoh"]
        assert "The Longest Night" in result["Pharaoh"]

    def test_returns_empty_dict_when_all_searched(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_SEARCHED)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        ws.cell(row=2, column=3, value=True)
        assert get_unsearched_artists(ws) == {}

    def test_skips_blank_rows(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE)
        ws = wb.active
        ws.cell(row=2, column=1, value=None)  # blank artist
        ws.cell(row=2, column=2, value="Something")
        assert get_unsearched_artists(ws) == {}

    def test_no_duplicate_releases_for_same_artist(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        ws.cell(row=3, column=1, value="Pharaoh")
        ws.cell(row=3, column=2, value="After the Fire")
        result = get_unsearched_artists(ws)
        assert result["Pharaoh"].count("After the Fire") == 1


# ---------------------------------------------------------------------------
# pick_random_artist
# ---------------------------------------------------------------------------

class TestPickRandomArtist:
    def test_returns_none_when_all_searched(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_SEARCHED)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        ws.cell(row=2, column=3, value=True)

        artist, releases = pick_random_artist(ws)
        assert artist is None
        assert releases is None

    def test_returns_unsearched_artist_and_releases(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")

        artist, releases = pick_random_artist(ws)
        assert artist == "Pharaoh"
        assert "After the Fire" in releases


# ---------------------------------------------------------------------------
# find_duplicate_rows
# ---------------------------------------------------------------------------

class TestFindDuplicateRows:
    def _make_ws(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_SEARCHED)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        ws.cell(row=3, column=1, value="Iced Earth")
        ws.cell(row=3, column=2, value="Framing Armageddon")
        return ws

    def test_returns_empty_when_no_duplicates(self):
        ws = self._make_ws()
        assert find_duplicate_rows(ws) == []

    def test_detects_duplicate_pair(self):
        ws = self._make_ws()
        ws.cell(row=4, column=1, value="Pharaoh")
        ws.cell(row=4, column=2, value="After the Fire")
        dupes = find_duplicate_rows(ws)
        assert len(dupes) == 1
        artist, release, rows = dupes[0]
        assert artist == "Pharaoh"
        assert release == "After the Fire"
        assert sorted(rows) == [2, 4]

    def test_reports_all_row_numbers_for_triplicate(self):
        ws = self._make_ws()
        ws.cell(row=4, column=1, value="Pharaoh")
        ws.cell(row=4, column=2, value="After the Fire")
        ws.cell(row=5, column=1, value="Pharaoh")
        ws.cell(row=5, column=2, value="After the Fire")
        artist, release, rows = find_duplicate_rows(ws)[0]
        assert sorted(rows) == [2, 4, 5]

    def test_does_not_flag_same_artist_different_release(self):
        ws = self._make_ws()
        ws.cell(row=4, column=1, value="Pharaoh")
        ws.cell(row=4, column=2, value="The Longest Night")
        assert find_duplicate_rows(ws) == []

    def test_result_is_sorted_by_artist_then_release(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE)
        ws = wb.active
        for r, (a, rel) in enumerate([
            ("Iced Earth", "Framing Armageddon"),
            ("Pharaoh", "After the Fire"),
            ("Iced Earth", "Framing Armageddon"),
            ("Pharaoh", "After the Fire"),
        ], start=2):
            ws.cell(row=r, column=1, value=a)
            ws.cell(row=r, column=2, value=rel)
        dupes = find_duplicate_rows(ws)
        assert [d[0] for d in dupes] == ["Iced Earth", "Pharaoh"]

    def test_sort_is_case_insensitive(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE)
        ws = wb.active
        for r, (a, rel) in enumerate([
            ("zz Top", "Eliminator"),
            ("Accept", "Balls to the Wall"),
            ("zz Top", "Eliminator"),
            ("Accept", "Balls to the Wall"),
        ], start=2):
            ws.cell(row=r, column=1, value=a)
            ws.cell(row=r, column=2, value=rel)
        dupes = find_duplicate_rows(ws)
        assert dupes[0][0] == "Accept"
        assert dupes[1][0] == "zz Top"


# ---------------------------------------------------------------------------
# update_release_row
# ---------------------------------------------------------------------------

class TestUpdateReleaseRow:
    def _make_ws(self):
        wb = _wb_with_headers(
            C_ARTIST, C_RELEASE, C_YEAR, C_GENRE,
            C_MA_ARTIST_URL, C_MA_RELEASE_URL, C_TYPE,
            C_SEARCHED, C_FOUND, C_REVIEW_FLAG
        )
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        ws.cell(row=2, column=3, value=2008)
        return ws

    def test_sets_searched_and_found(self):
        ws = self._make_ws()
        update_release_row(ws, "Pharaoh", "After the Fire", searched=True, found=True)
        assert _col_val(ws, 2, C_SEARCHED) is True
        assert _col_val(ws, 2, C_FOUND) is True

    def test_sets_optional_url_fields(self):
        ws = self._make_ws()
        update_release_row(
            ws, "Pharaoh", "After the Fire",
            searched=True, found=True,
            artist_url="http://ma.com/bands/Pharaoh/2801",
            release_url="http://ma.com/albums/Pharaoh/After/123",
            release_type="Full-length",
        )
        assert _col_val(ws, 2, C_MA_ARTIST_URL) == "http://ma.com/bands/Pharaoh/2801"
        assert _col_val(ws, 2, C_MA_RELEASE_URL) == "http://ma.com/albums/Pharaoh/After/123"
        assert _col_val(ws, 2, C_TYPE) == "Full-length"

    def test_does_not_set_optional_fields_when_none(self):
        ws = self._make_ws()
        update_release_row(ws, "Pharaoh", "After the Fire", searched=True, found=False)
        assert _col_val(ws, 2, C_MA_ARTIST_URL) is None
        assert _col_val(ws, 2, C_MA_RELEASE_URL) is None

    def test_sets_needs_review_flag(self):
        ws = self._make_ws()
        update_release_row(ws, "Pharaoh", "After the Fire",
                           searched=True, found=True, needs_review=True)
        assert _col_val(ws, 2, C_REVIEW_FLAG) is True

    def test_does_not_overwrite_existing_type(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_TYPE, C_SEARCHED, C_FOUND)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        ws.cell(row=2, column=3, value="EP")

        update_release_row(ws, "Pharaoh", "After the Fire",
                           searched=True, found=True, release_type="Full-length")
        assert _col_val(ws, 2, C_TYPE) == "EP"

    def test_does_not_overwrite_existing_year(self):
        ws = self._make_ws()  # row 2 already has year=2008
        update_release_row(ws, "Pharaoh", "After the Fire",
                           searched=True, found=True, year="1999")
        assert _col_val(ws, 2, C_YEAR) == 2008

    def test_writes_year_when_cell_is_empty(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_YEAR, C_SEARCHED, C_FOUND)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        # year cell intentionally left empty

        update_release_row(ws, "Pharaoh", "After the Fire",
                           searched=True, found=True, year="2008")
        assert _col_val(ws, 2, C_YEAR) == 2008

    def test_does_not_overwrite_existing_genre(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_GENRE, C_SEARCHED, C_FOUND)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        ws.cell(row=2, column=3, value="Power Metal")

        update_release_row(ws, "Pharaoh", "After the Fire",
                           searched=True, found=True, genre="Different Genre")
        assert _col_val(ws, 2, C_GENRE) == "Power Metal"

    def test_writes_genre_when_cell_is_empty(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_GENRE, C_SEARCHED, C_FOUND)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        # genre cell intentionally left empty

        update_release_row(ws, "Pharaoh", "After the Fire",
                           searched=True, found=True, genre="Power/Progressive Metal")
        assert _col_val(ws, 2, C_GENRE) == "Power/Progressive Metal"

    def test_does_not_modify_non_matching_rows(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_SEARCHED, C_FOUND)
        ws = wb.active
        ws.cell(row=2, column=1, value="Pharaoh")
        ws.cell(row=2, column=2, value="After the Fire")
        ws.cell(row=3, column=1, value="Pharaoh")
        ws.cell(row=3, column=2, value="The Longest Night")

        update_release_row(ws, "Pharaoh", "After the Fire", searched=True, found=True)
        # Row 3 (The Longest Night) should be untouched
        assert ws.cell(row=3, column=3).value is None

    def test_no_op_on_artist_not_in_sheet(self):
        ws = self._make_ws()
        # Should not raise
        update_release_row(ws, "Nonexistent", "Album", searched=True, found=False)

    def test_updates_unsearched_duplicate_when_later_row_is_already_searched(self):
        # Regression: _get_row_index used a dict with last-row-wins semantics.
        # If an already-searched duplicate appeared after the unsearched row,
        # update_release_row would write to the searched row (no-op) and leave
        # the unsearched duplicate permanently stuck.
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_SEARCHED, C_FOUND)
        ws = wb.active
        # Row 2: unsearched duplicate (the one that must be updated)
        ws.cell(row=2, column=1, value="Withering Surface")
        ws.cell(row=2, column=2, value="Exit Plan")
        # Row 3: already-searched duplicate (appears later — was the last-wins target)
        ws.cell(row=3, column=1, value="Withering Surface")
        ws.cell(row=3, column=2, value="Exit Plan")
        ws.cell(row=3, column=3, value=True)   # Searched
        ws.cell(row=3, column=4, value=True)   # Found

        update_release_row(ws, "Withering Surface", "Exit Plan", searched=True, found=False)

        assert ws.cell(row=2, column=3).value is True, (
            "Unsearched duplicate (row 2) was not marked — update went to the already-searched row instead"
        )

    def test_updates_all_unsearched_duplicates(self):
        # When every duplicate is unsearched, all of them must be marked so
        # none remain visible to get_unsearched_artists.
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_SEARCHED, C_FOUND)
        ws = wb.active
        ws.cell(row=2, column=1, value="Withering Surface")
        ws.cell(row=2, column=2, value="Exit Plan")
        ws.cell(row=3, column=1, value="Withering Surface")
        ws.cell(row=3, column=2, value="Exit Plan")

        update_release_row(ws, "Withering Surface", "Exit Plan", searched=True, found=False)

        assert ws.cell(row=2, column=3).value is True
        assert ws.cell(row=3, column=3).value is True


# ---------------------------------------------------------------------------
# _compute_disambiguations
# ---------------------------------------------------------------------------

class TestComputeDisambiguations:
    def _a(self, country="", location="", formed_in="", genre="", years_active=""):
        return {"country": country, "location": location, "formed_in": formed_in,
                "genre": genre, "years_active": years_active}

    def test_single_artist_needs_no_disambiguation(self):
        result = _compute_disambiguations([self._a(country="US")])
        assert result == [""]

    def test_empty_list_returns_empty(self):
        assert _compute_disambiguations([]) == []

    def test_different_countries_disambiguated_by_country_alone(self):
        us = self._a(country="United States", location="Philadelphia")
        sa = self._a(country="Saudi Arabia", location="Riyadh")
        result = _compute_disambiguations([us, sa])
        assert result == ["United States", "Saudi Arabia"]

    def test_same_country_requires_location(self):
        a = self._a(country="United States", location="Philadelphia")
        b = self._a(country="United States", location="New Jersey")
        result = _compute_disambiguations([a, b])
        assert result == ["United States | Philadelphia", "United States | New Jersey"]

    def test_unique_by_country_gets_no_extra_fields(self):
        sa = self._a(country="Saudi Arabia", location="Riyadh")
        us_a = self._a(country="United States", location="Philadelphia")
        us_b = self._a(country="United States", location="New Jersey")
        result = _compute_disambiguations([sa, us_a, us_b])
        assert result[0] == "Saudi Arabia"
        assert "Philadelphia" in result[1]
        assert "New Jersey" in result[2]

    def test_fully_identical_artists_return_same_string(self):
        a = self._a(country="US", location="NY", genre="Metal")
        b = self._a(country="US", location="NY", genre="Metal")
        result = _compute_disambiguations([a, b])
        assert result[0] == result[1]

    def test_skips_empty_field_values(self):
        a = self._a(country="Germany", location="")
        b = self._a(country="Sweden", location="")
        result = _compute_disambiguations([a, b])
        assert result == ["Germany", "Sweden"]


# ---------------------------------------------------------------------------
# update_artist_row
# ---------------------------------------------------------------------------

class TestUpdateArtistRow:
    def _make_artists_ws(self):
        wb = _wb_with_headers(
            COL_ARTISTS_ARTIST, COL_ARTISTS_DISAMBIG,
            COL_ARTISTS_LOCATION, COL_ARTISTS_MA_URL, COL_ARTISTS_MA_INFO
        )
        return wb.active

    def test_adds_new_artist_row(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        update_artist_row(ws, pharaoh_artist_dict)
        assert ws.max_row == 2
        assert _col_val(ws, 2, COL_ARTISTS_ARTIST) == "Pharaoh"

    def test_stores_url(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        update_artist_row(ws, pharaoh_artist_dict)
        assert _col_val(ws, 2, COL_ARTISTS_MA_URL) == pharaoh_artist_dict["url"]

    def test_stores_location_as_country_slash_location(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        update_artist_row(ws, pharaoh_artist_dict)
        location = _col_val(ws, 2, COL_ARTISTS_LOCATION)
        assert "United States" in location
        assert "Philadelphia, PA" in location

    def test_stores_json_info_excluding_discography(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        update_artist_row(ws, pharaoh_artist_dict)
        raw = _col_val(ws, 2, COL_ARTISTS_MA_INFO)
        data = json.loads(raw)
        assert "discography" not in data
        assert data["name"] == "Pharaoh"

    def test_skips_duplicate_url(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        update_artist_row(ws, pharaoh_artist_dict)
        update_artist_row(ws, pharaoh_artist_dict)  # second call
        assert ws.max_row == 2  # still only the header + one data row

    def test_single_artist_gets_no_disambiguation(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        update_artist_row(ws, pharaoh_artist_dict)
        assert _col_val(ws, 2, COL_ARTISTS_DISAMBIG) == ""

    def test_two_artists_different_countries_both_stored_with_country_disambig(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        other = dict(pharaoh_artist_dict)
        other["url"] = "https://www.metal-archives.com/bands/Pharaoh/9999"
        other["country"] = "Saudi Arabia"
        other["location"] = "Riyadh"
        update_artist_row(ws, pharaoh_artist_dict)
        update_artist_row(ws, other)
        assert ws.max_row == 3
        assert _col_val(ws, 2, COL_ARTISTS_DISAMBIG) == "United States"
        assert _col_val(ws, 3, COL_ARTISTS_DISAMBIG) == "Saudi Arabia"

    def test_retroactively_updates_first_artist_when_duplicate_name_added(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        update_artist_row(ws, pharaoh_artist_dict)
        # First artist has no disambiguation yet
        assert _col_val(ws, 2, COL_ARTISTS_DISAMBIG) == ""

        other = dict(pharaoh_artist_dict)
        other["url"] = "https://www.metal-archives.com/bands/Pharaoh/9999"
        other["country"] = "Germany"
        update_artist_row(ws, other)
        # First artist should now have its country as disambiguation
        assert _col_val(ws, 2, COL_ARTISTS_DISAMBIG) == "United States"

    def test_same_country_requires_location_to_disambiguate(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        other = dict(pharaoh_artist_dict)
        other["url"] = "https://www.metal-archives.com/bands/Pharaoh/9999"
        other["location"] = "Somerville, NJ"
        update_artist_row(ws, pharaoh_artist_dict)
        update_artist_row(ws, other)
        d1 = _col_val(ws, 2, COL_ARTISTS_DISAMBIG)
        d2 = _col_val(ws, 3, COL_ARTISTS_DISAMBIG)
        assert "Philadelphia" in d1
        assert "Somerville" in d2
        assert d1 != d2

    def test_fully_identical_artists_get_numeric_suffix(self, pharaoh_artist_dict):
        ws = self._make_artists_ws()
        a2 = dict(pharaoh_artist_dict)
        a2["url"] = "https://www.metal-archives.com/bands/Pharaoh/9999"
        update_artist_row(ws, pharaoh_artist_dict)
        update_artist_row(ws, a2)
        names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert any("(2)" in (n or "") for n in names)


# ---------------------------------------------------------------------------
# add_review_entry
# ---------------------------------------------------------------------------

class TestAddReviewEntry:
    def _make_review_ws(self):
        wb = _wb_with_headers(
            COL_REVIEW_ARTIST, COL_REVIEW_RELEASE, COL_REVIEW_ISSUE, COL_REVIEW_MA_URLS
        )
        return wb.active

    def test_adds_row_with_all_fields(self):
        ws = self._make_review_ws()
        add_review_entry(ws, "Pharaoh", "After the Fire", "Ambiguous match",
                         ["http://ma.com/1", "http://ma.com/2"])
        assert ws.max_row == 2
        assert _col_val(ws, 2, COL_REVIEW_ARTIST) == "Pharaoh"
        assert _col_val(ws, 2, COL_REVIEW_RELEASE) == "After the Fire"
        assert _col_val(ws, 2, COL_REVIEW_ISSUE) == "Ambiguous match"

    def test_joins_multiple_urls_with_pipe(self):
        ws = self._make_review_ws()
        add_review_entry(ws, "Band", "Album", "Issue",
                         ["http://url1", "http://url2"])
        urls = _col_val(ws, 2, COL_REVIEW_MA_URLS)
        assert "http://url1" in urls
        assert "http://url2" in urls
        assert " | " in urls

    def test_multiple_entries_appended(self):
        ws = self._make_review_ws()
        add_review_entry(ws, "Band A", "Album A", "Issue", ["http://a"])
        add_review_entry(ws, "Band B", "Album B", "Issue", ["http://b"])
        assert ws.max_row == 3

    def test_does_not_add_duplicate_entries(self):
        ws = self._make_review_ws()
        add_review_entry(ws, "Pharaoh", "After the Fire", "Ambiguous match", ["http://a"])
        add_review_entry(ws, "Pharaoh", "After the Fire", "Ambiguous match", ["http://a"])
        assert ws.max_row == 2

    def test_different_issue_for_same_release_is_added(self):
        ws = self._make_review_ws()
        add_review_entry(ws, "Pharaoh", "After the Fire", "Ambiguous match", ["http://a"])
        add_review_entry(ws, "Pharaoh", "After the Fire", "Partial match located", ["http://b"])
        assert ws.max_row == 3


# ---------------------------------------------------------------------------
# expand_st_titles
# ---------------------------------------------------------------------------

class TestExpandStTitles:
    def _make_ws(self, release_title="s/t"):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_SEARCHED, C_FOUND)
        ws_collection = wb.active
        ws_collection.cell(row=2, column=1, value="Pharaoh")
        ws_collection.cell(row=2, column=2, value=release_title)
        wb.create_sheet("Review")
        ws_review = wb["Review"]
        for i, col in enumerate(
            [COL_REVIEW_ARTIST, COL_REVIEW_RELEASE, COL_REVIEW_ISSUE, COL_REVIEW_MA_URLS],
            start=1,
        ):
            ws_review.cell(row=1, column=i, value=col)
        return ws_collection, ws_review

    def test_replaces_st_with_artist_name_in_cell(self):
        ws_collection, ws_review = self._make_ws("s/t")
        expand_st_titles(ws_collection, ws_review, "Pharaoh", ["s/t"])
        assert ws_collection.cell(row=2, column=2).value == "Pharaoh"

    def test_returns_artist_name_in_place_of_st(self):
        ws_collection, ws_review = self._make_ws("s/t")
        result = expand_st_titles(ws_collection, ws_review, "Pharaoh", ["s/t"])
        assert result == ["Pharaoh"]

    def test_adds_review_entry_for_st_substitution(self):
        ws_collection, ws_review = self._make_ws("s/t")
        expand_st_titles(ws_collection, ws_review, "Pharaoh", ["s/t"])
        assert ws_review.max_row == 2
        assert ws_review.cell(row=2, column=3).value == "s/t changed to artist name"

    def test_case_insensitive_match(self):
        ws_collection, ws_review = self._make_ws("S/T")
        result = expand_st_titles(ws_collection, ws_review, "Pharaoh", ["S/T"])
        assert result == ["Pharaoh"]
        assert ws_review.max_row == 2

    def test_non_st_titles_pass_through_unchanged(self):
        ws_collection, ws_review = self._make_ws("After the Fire")
        result = expand_st_titles(ws_collection, ws_review, "Pharaoh", ["After the Fire"])
        assert result == ["After the Fire"]
        assert ws_review.max_row == 1  # no review entry added

    def test_mixed_list_only_replaces_st(self):
        ws_collection, ws_review = self._make_ws()
        ws_collection.cell(row=3, column=1, value="Pharaoh")
        ws_collection.cell(row=3, column=2, value="The Longest Night")
        result = expand_st_titles(
            ws_collection, ws_review, "Pharaoh", ["s/t", "The Longest Night"]
        )
        assert result == ["Pharaoh", "The Longest Night"]
        assert ws_collection.cell(row=2, column=2).value == "Pharaoh"
        assert ws_collection.cell(row=3, column=2).value == "The Longest Night"

    def test_only_updates_matching_artist_rows(self):
        ws_collection, ws_review = self._make_ws("s/t")
        ws_collection.cell(row=3, column=1, value="Other Band")
        ws_collection.cell(row=3, column=2, value="s/t")
        expand_st_titles(ws_collection, ws_review, "Pharaoh", ["s/t"])
        # Other Band's row should be untouched
        assert ws_collection.cell(row=3, column=2).value == "s/t"

    def test_update_release_row_works_after_st_rename_with_stale_cache(self):
        # Regression: expand_st_titles modifies the Release cell without
        # invalidating _row_index_cache. A second artist with s/t could have
        # its cache entry built under ("Artist", "s/t") before the rename,
        # so the subsequent update_release_row("Artist", "Artist") lookup
        # would miss and silently skip writing the Searched flag — causing
        # the artist to loop forever as "unsearched".
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_SEARCHED, C_FOUND)
        ws_collection = wb.active
        # Artist 1 (non-s/t) — processed first, which populates the cache
        ws_collection.cell(row=2, column=1, value="Pharaoh")
        ws_collection.cell(row=2, column=2, value="After the Fire")
        # Artist 2 (s/t) — cache is built before expand_st_titles renames it
        ws_collection.cell(row=3, column=1, value="Black Funeral")
        ws_collection.cell(row=3, column=2, value="s/t")

        wb.create_sheet("Review")
        ws_review = wb["Review"]
        for i, col in enumerate(
            [COL_REVIEW_ARTIST, COL_REVIEW_RELEASE, COL_REVIEW_ISSUE, COL_REVIEW_MA_URLS],
            start=1,
        ):
            ws_review.cell(row=1, column=i, value=col)

        # Trigger cache build (simulates first artist's update_release_row)
        update_release_row(ws_collection, "Pharaoh", "After the Fire",
                           searched=True, found=True)

        # Now expand s/t for artist 2 — this renames the cell and must clear the cache
        expand_st_titles(ws_collection, ws_review, "Black Funeral", ["s/t"])

        # update_release_row must find the row even though cache was stale
        update_release_row(ws_collection, "Black Funeral", "Black Funeral",
                           searched=True, found=False)

        searched = _col_val(ws_collection, 3, C_SEARCHED)
        assert searched is True, (
            "Searched flag was not written — expand_st_titles likely failed to "
            "invalidate _row_index_cache after renaming the s/t cell"
        )


# ---------------------------------------------------------------------------
# add_not_found_entry
# ---------------------------------------------------------------------------

class TestAddNotFoundEntry:
    def _make_ws(self):
        wb = _wb_with_headers(COL_NOT_FOUND_ARTIST, COL_NOT_FOUND_RELEASE)
        return wb.active

    def test_adds_artist_and_release(self):
        ws = self._make_ws()
        add_not_found_entry(ws, "Pharaoh", "After the Fire")
        assert _col_val(ws, 2, COL_NOT_FOUND_ARTIST) == "Pharaoh"
        assert _col_val(ws, 2, COL_NOT_FOUND_RELEASE) == "After the Fire"

    def test_does_not_add_duplicate_entries(self):
        ws = self._make_ws()
        add_not_found_entry(ws, "Pharaoh", "After the Fire")
        add_not_found_entry(ws, "Pharaoh", "After the Fire")
        assert ws.max_row == 2  # header + one data row

    def test_different_releases_for_same_artist_both_added(self):
        ws = self._make_ws()
        add_not_found_entry(ws, "Pharaoh", "After the Fire")
        add_not_found_entry(ws, "Pharaoh", "The Longest Night")
        assert ws.max_row == 3


# ---------------------------------------------------------------------------
# save_workbook
# ---------------------------------------------------------------------------

class TestSaveWorkbook:
    def test_saves_workbook_to_path(self, tmp_path):
        wb = Workbook()
        wb.active["A1"] = "test"
        path = str(tmp_path / "output.xlsx")
        save_workbook(wb, path)
        assert (tmp_path / "output.xlsx").exists()

    def test_no_tmp_file_left_after_successful_save(self, tmp_path):
        wb = Workbook()
        wb.active["A1"] = "test"
        path = str(tmp_path / "output.xlsx")
        save_workbook(wb, path)
        assert not (tmp_path / "output.tmp.xlsx").exists()

    def test_original_file_preserved_when_write_fails(self, tmp_path):
        # Pre-write a valid spreadsheet so there is an "original" to preserve.
        from openpyxl import load_workbook
        original = Workbook()
        original.active["A1"] = "original"
        path = str(tmp_path / "output.xlsx")
        original.save(path)

        from unittest.mock import MagicMock
        bad_wb = MagicMock()
        bad_wb.save.side_effect = [PermissionError("locked"), None]  # primary fails, backup succeeds
        save_workbook(bad_wb, path)

        # Original file must still be intact and readable.
        loaded = load_workbook(path)
        assert loaded.active["A1"].value == "original"

    def test_tmp_file_removed_on_write_failure(self, tmp_path):
        from unittest.mock import MagicMock
        wb = MagicMock()
        wb.save.side_effect = [PermissionError("locked"), None]
        path = str(tmp_path / "output.xlsx")
        save_workbook(wb, path)
        assert not (tmp_path / "output.tmp.xlsx").exists()

    def test_uses_backup_path_on_primary_failure(self, tmp_path):
        from unittest.mock import MagicMock
        wb = MagicMock()
        wb.save.side_effect = [PermissionError("locked"), None]
        path = str(tmp_path / "output.xlsx")
        save_workbook(wb, path)
        # save called twice: once for tmp (primary), once for backup
        assert wb.save.call_count == 2
        backup_path = wb.save.call_args_list[1][0][0]
        assert "backup" in backup_path

    def test_raises_when_both_primary_and_backup_fail(self, tmp_path):
        from unittest.mock import MagicMock
        wb = MagicMock()
        wb.save.side_effect = PermissionError("always locked")
        path = str(tmp_path / "output.xlsx")
        with pytest.raises(PermissionError):
            save_workbook(wb, path)


# ---------------------------------------------------------------------------
# update_stats_sheet
# ---------------------------------------------------------------------------

class TestUpdateStatsSheet:
    def _make_full_wb(self):
        wb = _wb_with_headers(C_ARTIST, C_RELEASE, C_YEAR, C_GENRE, C_TYPE,
                              C_SEARCHED, C_FOUND, C_REVIEW_FLAG)
        ensure_collection_sheet(wb)
        ensure_artists_sheet(wb)
        ensure_review_sheet(wb)
        ensure_not_found_sheet(wb)
        return wb

    def test_creates_statistics_sheet(self):
        wb = self._make_full_wb()
        update_stats_sheet(wb)
        assert STATS_SHEET in wb.sheetnames

    def test_does_nothing_without_collection_sheet(self):
        wb = Workbook()
        wb.active.title = "NotCollection"
        update_stats_sheet(wb)
        assert STATS_SHEET not in wb.sheetnames

    def test_idempotent_does_not_duplicate_sheet(self):
        wb = self._make_full_wb()
        update_stats_sheet(wb)
        update_stats_sheet(wb)
        assert wb.sheetnames.count(STATS_SHEET) == 1

    def test_expected_labels_present(self):
        wb = self._make_full_wb()
        update_stats_sheet(wb)
        ws = wb[STATS_SHEET]
        labels = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
        for expected in [
            "Collection", "Release Types", "Metadata", "Other Sheets",
            "Total Releases", "Unique Artists", "Searched", "Found",
            "Completion", "Found Rate", "With Year", "With Genre", "Unique Genres",
            "Artists Processed", "Requiring Review", "Not Found", "Last Updated",
        ]:
            assert expected in labels, f"Missing label: {expected}"

    def test_completion_and_found_rate_have_pct_format(self):
        wb = self._make_full_wb()
        update_stats_sheet(wb)
        ws = wb[STATS_SHEET]
        pct_rows = [
            r for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=1).value in ("Completion", "Found Rate")
        ]
        assert len(pct_rows) == 2
        for row in pct_rows:
            assert "%" in ws.cell(row=row, column=2).number_format

    def test_value_formulas_reference_collection(self):
        wb = self._make_full_wb()
        update_stats_sheet(wb)
        ws = wb[STATS_SHEET]
        formulas = [
            ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(row=r, column=2).value, str)
            and ws.cell(row=r, column=2).value.startswith("=")
        ]
        assert any("Collection" in f for f in formulas)

    def test_last_updated_is_static_string(self):
        from datetime import datetime as _dt
        wb = self._make_full_wb()
        update_stats_sheet(wb)
        ws = wb[STATS_SHEET]
        val = next(
            (ws.cell(row=r, column=2).value
             for r in range(1, ws.max_row + 1)
             if ws.cell(row=r, column=1).value == "Last Updated"),
            None,
        )
        assert val is not None
        _dt.strptime(val, "%Y-%m-%d %H:%M:%S")  # raises if format is wrong
