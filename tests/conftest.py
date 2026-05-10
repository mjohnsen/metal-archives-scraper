import pytest
from openpyxl import Workbook

import metal_archives_scraper.spreadsheet as _ss


@pytest.fixture(autouse=True)
def _clear_spreadsheet_caches():
    """Prevent id(worksheet) collisions between tests from returning stale cached data."""
    _ss._col_map_cache.clear()
    _ss._row_index_cache.clear()
    yield
    _ss._col_map_cache.clear()
    _ss._row_index_cache.clear()

from metal_archives_scraper.spreadsheet import (
    C_ARTIST, C_FOUND, C_GENRE, C_MA_ARTIST_URL, C_MA_RELEASE_URL,
    C_REVIEW_FLAG, C_SEARCHED, C_TYPE, C_YEAR, C_RELEASE,
    COL_ARTISTS_ARTIST, COL_ARTISTS_DISAMBIG, COL_ARTISTS_LOCATION,
    COL_ARTISTS_MA_INFO, COL_ARTISTS_MA_URL,
    COL_NOT_FOUND_ARTIST, COL_NOT_FOUND_RELEASE,
    COL_REVIEW_ARTIST, COL_REVIEW_ISSUE, COL_REVIEW_MA_URLS, COL_REVIEW_RELEASE,
)

COLLECTION_HEADERS = [
    C_ARTIST, C_RELEASE, C_YEAR, C_GENRE,
    C_MA_ARTIST_URL, C_MA_RELEASE_URL, C_TYPE,
    C_SEARCHED, C_FOUND, C_REVIEW_FLAG,
]

ARTISTS_HEADERS = [
    COL_ARTISTS_ARTIST, COL_ARTISTS_DISAMBIG,
    COL_ARTISTS_LOCATION, COL_ARTISTS_MA_URL, COL_ARTISTS_MA_INFO,
]

REVIEW_HEADERS = [
    COL_REVIEW_ARTIST, COL_REVIEW_RELEASE, COL_REVIEW_ISSUE, COL_REVIEW_MA_URLS,
]

NOT_FOUND_HEADERS = [COL_NOT_FOUND_ARTIST, COL_NOT_FOUND_RELEASE]


def _make_sheet(wb, title, headers):
    ws = wb.create_sheet(title)
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    return ws


@pytest.fixture
def minimal_wb():
    """Workbook with only Artist and Release columns (minimum viable)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value=C_ARTIST)
    ws.cell(row=1, column=2, value=C_RELEASE)
    return wb


@pytest.fixture
def full_wb():
    """Workbook with all Collection columns and representative data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"
    for col, name in enumerate(COLLECTION_HEADERS, start=1):
        ws.cell(row=1, column=col, value=name)

    rows = [
        # (Artist, Release, Year, Genre, MAartistURL, MAreleaseURL, Type, Searched, Found, Review)
        ("Pharaoh", "After the Fire", 2008, "Power Metal", None, None, None, None, None, None),
        ("Pharaoh", "The Longest Night", 2006, "Power Metal", None, None, None, None, None, None),
        ("Iced Earth", "Framing Armageddon", 2007, "Heavy Metal", None, None, None, True, True, None),
        ("Iced Earth", "The Crucible of Man", 2008, "Heavy Metal", None, None, None, None, None, None),
    ]
    for row_num, row_data in enumerate(rows, start=2):
        for col, val in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col, value=val)

    _make_sheet(wb, "Artists", ARTISTS_HEADERS)
    _make_sheet(wb, "Review", REVIEW_HEADERS)
    _make_sheet(wb, "Not Found", NOT_FOUND_HEADERS)
    return wb


@pytest.fixture
def pharaoh_artist_dict():
    return {
        "name": "Pharaoh",
        "url": "https://www.metal-archives.com/bands/Pharaoh/2801",
        "band_id": "2801",
        "country": "United States",
        "location": "Philadelphia, PA",
        "status": "Active",
        "formed_in": "2000",
        "genre": "Power/Progressive Metal",
        "themes": "Philosophy",
        "label": "Cruz Del Sur Music",
        "years_active": "2000-present",
        "discography": [
            {
                "title": "After the Fire",
                "url": "https://www.metal-archives.com/albums/Pharaoh/After_the_Fire/123",
                "type": "Full-length",
                "year": "2008",
            },
            {
                "title": "The Longest Night",
                "url": "https://www.metal-archives.com/albums/Pharaoh/The_Longest_Night/456",
                "type": "Full-length",
                "year": "2006",
            },
            {
                "title": "Be Gone",
                "url": "https://www.metal-archives.com/albums/Pharaoh/Be_Gone/789",
                "type": "Full-length",
                "year": "2003",
            },
        ],
    }
