from __future__ import annotations

import json
import logging
import os
import random
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Column name constants
C_ARTIST = "Artist"
C_RELEASE = "Release"
C_YEAR = "Year"
C_GENRE = "Genre"
C_MA_ARTIST_URL = "Metal Archives Artist URL"
C_MA_RELEASE_URL = "Metal Archives Release URL"
C_TYPE = "Type"
C_SEARCHED = "Searched"
C_FOUND = "Found"
C_REVIEW_FLAG = "Needs Review"

COL_ARTISTS_ARTIST = "Artist"
COL_ARTISTS_DISAMBIG = "Disambiguation"
COL_ARTISTS_LOCATION = "Location"
COL_ARTISTS_MA_URL = "Metal Archives URL"
COL_ARTISTS_MA_INFO = "Metal Archives Information"

COL_REVIEW_ARTIST = "Artist"
COL_REVIEW_RELEASE = "Release"
COL_REVIEW_ISSUE = "Issue"
COL_REVIEW_MA_URLS = "Metal Archives URLs"

COL_NOT_FOUND_ARTIST = "Artist"
COL_NOT_FOUND_RELEASE = "Release"

STATS_SHEET = "Statistics"
_STATS_MAX_ROWS = 50000

# Caches keyed by id(worksheet). Cleared between test runs via conftest fixture.
# In production a worksheet object lives for the entire session, so these are
# populated once and never stale.
_col_map_cache: dict[int, dict[str, int]] = {}
_row_index_cache: dict[int, dict[tuple[str, str], int]] = {}

COLLECTION_REQUIRED = [C_ARTIST, C_RELEASE]
# Inserted in this exact order immediately after Artist/Release
COLLECTION_ORDERED = [C_YEAR, C_GENRE, C_TYPE]
# Appended (in any order) after the ordered group
COLLECTION_ADDED = [
    C_MA_ARTIST_URL,
    C_MA_RELEASE_URL,
    C_SEARCHED,
    C_FOUND,
    C_REVIEW_FLAG,
]


def _col_index(ws: Worksheet, name: str) -> int | None:
    for cell in ws[1]:
        if cell.value == name:
            return cell.column
    return None


def _ensure_column(ws: Worksheet, name: str) -> int:
    idx = _col_index(ws, name)
    if idx is None:
        idx = ws.max_column + 1
        ws.cell(row=1, column=idx, value=name)
        _col_map_cache.pop(id(ws), None)  # invalidate after adding a column
    return idx


def _get_col_map(ws: Worksheet) -> dict[str, int]:
    ws_id = id(ws)
    if ws_id not in _col_map_cache:
        _col_map_cache[ws_id] = {cell.value: cell.column for cell in ws[1] if cell.value}
    return _col_map_cache[ws_id]


def _get_row_index(ws: Worksheet) -> dict[tuple[str, str], int]:
    ws_id = id(ws)
    if ws_id not in _row_index_cache:
        col_map = _get_col_map(ws)
        artist_col = col_map.get(C_ARTIST)
        release_col = col_map.get(C_RELEASE)
        index: dict[tuple[str, str], int] = {}
        if artist_col and release_col:
            for row in ws.iter_rows(min_row=2):
                artist = row[artist_col - 1].value
                release = row[release_col - 1].value
                if artist and release:
                    index[(str(artist), str(release))] = row[0].row
        _row_index_cache[ws_id] = index
    return _row_index_cache[ws_id]


def open_workbook(path: str) -> Workbook:
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    headers = {cell.value for cell in ws[1]}
    missing = [c for c in COLLECTION_REQUIRED if c not in headers]
    if missing:
        raise ValueError(f"Collection spreadsheet is missing required columns: {missing}")
    return wb


def _ensure_ordered_columns(ws: Worksheet, column_names: list[str]) -> None:
    """Insert missing columns in order, each placed right after its predecessor.

    The first column in column_names is positioned immediately after Release.
    Columns that already exist are not moved; only absent ones are inserted at
    the correct position relative to their neighbours in column_names.
    """
    col_map = _get_col_map(ws)
    insert_after = col_map.get(C_RELEASE, 2)

    for name in column_names:
        if name not in col_map:
            ws.insert_cols(insert_after + 1)
            ws.cell(row=1, column=insert_after + 1, value=name)
            _col_map_cache.pop(id(ws), None)
            _row_index_cache.pop(id(ws), None)
            col_map = _get_col_map(ws)
        insert_after = col_map[name]


def ensure_collection_sheet(wb: Workbook) -> Worksheet:
    ws = wb.worksheets[0]
    if ws.title != "Collection":
        ws.title = "Collection"

    _ensure_ordered_columns(ws, COLLECTION_ORDERED)

    existing_headers = {cell.value for cell in ws[1]}
    for col_name in COLLECTION_ADDED:
        if col_name not in existing_headers:
            _ensure_column(ws, col_name)

    return ws


def ensure_artists_sheet(wb: Workbook) -> Worksheet:
    if "Artists" not in wb.sheetnames:
        ws = wb.create_sheet("Artists")
        for col_name in [
            COL_ARTISTS_ARTIST,
            COL_ARTISTS_DISAMBIG,
            COL_ARTISTS_LOCATION,
            COL_ARTISTS_MA_URL,
            COL_ARTISTS_MA_INFO,
        ]:
            ws.cell(row=1, column=ws.max_column or 1, value=col_name)
            # max_column returns None on empty sheet first time
        # Redo properly
        ws.delete_rows(1, ws.max_row)
        for i, col_name in enumerate(
            [
                COL_ARTISTS_ARTIST,
                COL_ARTISTS_DISAMBIG,
                COL_ARTISTS_LOCATION,
                COL_ARTISTS_MA_URL,
                COL_ARTISTS_MA_INFO,
            ],
            start=1,
        ):
            ws.cell(row=1, column=i, value=col_name)
    return wb["Artists"]


def ensure_review_sheet(wb: Workbook) -> Worksheet:
    if "Review" not in wb.sheetnames:
        ws = wb.create_sheet("Review")
        for i, col_name in enumerate(
            [COL_REVIEW_ARTIST, COL_REVIEW_RELEASE, COL_REVIEW_ISSUE, COL_REVIEW_MA_URLS],
            start=1,
        ):
            ws.cell(row=1, column=i, value=col_name)
    return wb["Review"]


def ensure_not_found_sheet(wb: Workbook) -> Worksheet:
    if "Not Found" not in wb.sheetnames:
        ws = wb.create_sheet("Not Found")
        for i, col_name in enumerate(
            [COL_NOT_FOUND_ARTIST, COL_NOT_FOUND_RELEASE],
            start=1,
        ):
            ws.cell(row=1, column=i, value=col_name)
    return wb["Not Found"]


def get_unsearched_artists(ws: Worksheet) -> dict[str, list[str]]:
    col_map = _get_col_map(ws)
    artist_col = col_map.get(C_ARTIST)
    release_col = col_map.get(C_RELEASE)
    searched_col = col_map.get(C_SEARCHED)

    result: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not artist_col or not release_col:
            break
        artist = row[artist_col - 1]
        release = row[release_col - 1]
        searched = row[searched_col - 1] if searched_col else None

        if not artist or not release:
            continue
        if searched:
            continue

        result.setdefault(str(artist), [])
        if str(release) not in result[str(artist)]:
            result[str(artist)].append(str(release))

    return result


def pick_random_artist(ws: Worksheet) -> tuple[str, list[str]] | tuple[None, None]:
    unsearched = get_unsearched_artists(ws)
    if not unsearched:
        return None, None
    artist = random.choice(list(unsearched.keys()))
    return artist, unsearched[artist]


def update_release_row(
    ws: Worksheet,
    artist_name: str,
    release_title: str,
    searched: bool,
    found: bool,
    artist_url: str | None = None,
    release_url: str | None = None,
    release_type: str | None = None,
    year: str | None = None,
    genre: str | None = None,
    needs_review: bool = False,
):
    col_map = _get_col_map(ws)
    row_num = _get_row_index(ws).get((artist_name, release_title))
    if row_num is None:
        return

    def _set(col_name, value):
        col = col_map.get(col_name)
        if col and value is not None:
            ws.cell(row=row_num, column=col, value=value)

    _set(C_SEARCHED, searched)
    _set(C_FOUND, found)
    if artist_url:
        _set(C_MA_ARTIST_URL, artist_url)
    if release_url:
        _set(C_MA_RELEASE_URL, release_url)
    if needs_review:
        _set(C_REVIEW_FLAG, True)

    # Only populate year/genre/type when the cell is currently empty
    def _set_if_empty(col_name, value):
        col = col_map.get(col_name)
        if col and value:
            cell = ws.cell(row=row_num, column=col)
            if not cell.value:
                cell.value = value

    year_val = year
    if year:
        try:
            year_val = int(year)
        except (ValueError, TypeError):
            pass
    _set_if_empty(C_YEAR, year_val)
    _set_if_empty(C_GENRE, genre)
    _set_if_empty(C_TYPE, release_type)


_DISAMBIG_FIELDS = ("country", "location", "formed_in", "genre", "years_active")


def _compute_disambiguations(artists_info: list) -> list:
    """
    Given a list of artist dicts sharing a name, return the minimal list of
    disambiguation strings that makes every entry unique.

    Fields are tried in order until all artists can be told apart.  An artist
    stops receiving new fields once its string is already unique.  If all fields
    are exhausted and two artists are still identical, their strings will match
    (caller should fall back to a numeric suffix in that case).
    """
    n = len(artists_info)
    if n <= 1:
        return [""] * n

    disambigs = [""] * n
    needs_more = list(range(n))

    for field in _DISAMBIG_FIELDS:
        if not needs_more:
            break
        for j in needs_more:
            val = artists_info[j].get(field, "").strip()
            if val:
                disambigs[j] = f"{disambigs[j]} | {val}" if disambigs[j] else val
        needs_more = [
            j for j in needs_more
            if sum(1 for k in range(n) if disambigs[k] == disambigs[j]) > 1
        ]

    return disambigs


def update_artist_row(ws_artists: Worksheet, artist_dict: dict):
    col_map = _get_col_map(ws_artists)
    url_col = col_map.get(COL_ARTISTS_MA_URL)
    name_col = col_map.get(COL_ARTISTS_ARTIST)
    disambig_col = col_map.get(COL_ARTISTS_DISAMBIG)
    info_col = col_map.get(COL_ARTISTS_MA_INFO)

    target_url = artist_dict.get("url", "")
    target_name = artist_dict.get("name", "")

    # Skip if already recorded
    for row in ws_artists.iter_rows(min_row=2, values_only=True):
        existing_url = row[url_col - 1] if url_col else None
        if existing_url == target_url:
            return

    # Collect existing rows with the same artist name
    same_name_rows = []  # list of (row_num, info_dict)
    for row in ws_artists.iter_rows(min_row=2):
        name_cell = row[name_col - 1] if name_col else None
        info_cell = row[info_col - 1] if info_col else None
        if name_cell and name_cell.value == target_name:
            info_dict = {}
            if info_cell and info_cell.value:
                try:
                    info_dict = json.loads(info_cell.value)
                except Exception:
                    pass
            same_name_rows.append((row[0].row, info_dict))

    # Compute minimal disambiguations for all same-name artists + the new one
    all_infos = [r[1] for r in same_name_rows] + [artist_dict]
    disambigs = _compute_disambiguations(all_infos)

    # Retroactively update disambiguation cells for existing same-name rows
    for i, (row_num, _) in enumerate(same_name_rows):
        if disambig_col:
            ws_artists.cell(row=row_num, column=disambig_col, value=disambigs[i])

    new_disambig = disambigs[-1]

    # Fallback: if disambiguation is still not unique, add a numeric suffix to the display name
    existing_disambigs = [disambigs[i] for i in range(len(same_name_rows))]
    display_name = target_name
    if new_disambig in existing_disambigs:
        suffix = sum(1 for d in existing_disambigs if d == new_disambig) + 1
        display_name = f"{target_name} ({suffix})"

    new_row = ws_artists.max_row + 1
    ma_info = json.dumps(
        {k: v for k, v in artist_dict.items() if k != "discography"},
        ensure_ascii=False,
    )

    def _set(col_name, value):
        col = col_map.get(col_name)
        if col:
            ws_artists.cell(row=new_row, column=col, value=value)

    _set(COL_ARTISTS_ARTIST, display_name)
    _set(COL_ARTISTS_DISAMBIG, new_disambig)
    _set(COL_ARTISTS_LOCATION, f"{artist_dict.get('country', '')} / {artist_dict.get('location', '')}".strip(" /"))
    _set(COL_ARTISTS_MA_URL, target_url)
    _set(COL_ARTISTS_MA_INFO, ma_info)


def expand_st_titles(
    ws_collection: Worksheet,
    ws_review: Worksheet,
    artist_name: str,
    release_titles: list[str],
) -> list[str]:
    """Replace any 's/t' release title with the artist name in the spreadsheet.

    Returns the updated list of titles.  A review entry is added for each
    substitution so the change is visible to the user.
    """
    col_map = _get_col_map(ws_collection)
    artist_col = col_map.get(C_ARTIST)
    release_col = col_map.get(C_RELEASE)

    updated: list[str] = []
    for title in release_titles:
        if title.strip().lower() == "s/t":
            for row in ws_collection.iter_rows(min_row=2):
                row_artist = row[artist_col - 1].value if artist_col else None
                row_release = row[release_col - 1].value if release_col else None
                if (
                    str(row_artist) == artist_name
                    and str(row_release).strip().lower() == "s/t"
                ):
                    row[release_col - 1].value = artist_name
            add_review_entry(
                ws_review, artist_name, artist_name,
                "s/t changed to artist name", [],
            )
            updated.append(artist_name)
        else:
            updated.append(title)
    return updated


def add_review_entry(ws_review: Worksheet, artist_name: str, release_title: str, issue: str, ma_urls: list[str]):
    col_map = _get_col_map(ws_review)
    artist_col = col_map.get(COL_REVIEW_ARTIST)
    release_col = col_map.get(COL_REVIEW_RELEASE)
    issue_col = col_map.get(COL_REVIEW_ISSUE)
    for row in ws_review.iter_rows(min_row=2, values_only=True):
        if (
            row[artist_col - 1] == artist_name
            and row[release_col - 1] == release_title
            and row[issue_col - 1] == issue
        ):
            return
    new_row = ws_review.max_row + 1

    def _set(col_name, value):
        col = col_map.get(col_name)
        if col:
            ws_review.cell(row=new_row, column=col, value=value)

    _set(COL_REVIEW_ARTIST, artist_name)
    _set(COL_REVIEW_RELEASE, release_title)
    _set(COL_REVIEW_ISSUE, issue)
    _set(COL_REVIEW_MA_URLS, " | ".join(ma_urls))


def add_not_found_entry(ws_not_found: Worksheet, artist_name: str, release_title: str):
    col_map = _get_col_map(ws_not_found)
    artist_col = col_map.get(COL_NOT_FOUND_ARTIST)
    release_col = col_map.get(COL_NOT_FOUND_RELEASE)
    for row in ws_not_found.iter_rows(min_row=2, values_only=True):
        if row[artist_col - 1] == artist_name and row[release_col - 1] == release_title:
            return

    new_row = ws_not_found.max_row + 1

    def _set(col_name, value):
        col = col_map.get(col_name)
        if col:
            ws_not_found.cell(row=new_row, column=col, value=value)

    _set(COL_NOT_FOUND_ARTIST, artist_name)
    _set(COL_NOT_FOUND_RELEASE, release_title)


def update_stats_sheet(wb: Workbook) -> None:
    """Rebuild the Statistics sheet with Excel 365 formulas derived from the other sheets."""
    if "Collection" not in wb.sheetnames:
        return

    ws_col = wb["Collection"]
    col = _get_col_map(ws_col)

    def _cl(c):
        n = col.get(c)
        return get_column_letter(n) if n else None

    a = _cl(C_ARTIST)
    r = _cl(C_RELEASE)
    y = _cl(C_YEAR)
    g = _cl(C_GENRE)
    t = _cl(C_TYPE)
    s = _cl(C_SEARCHED)
    f = _cl(C_FOUND)
    N = _STATS_MAX_ROWS

    def _unique_count(letter):
        rng = f"Collection!{letter}2:{letter}{N}"
        return f'=IFERROR(SUMPRODUCT(({rng}<>"")/COUNTIF({rng},{rng}&"")),0)'

    def _countif_col(letter, value):
        return f"=COUNTIF(Collection!{letter}:{letter},{value})" if letter else 0

    def _counta_col(letter):
        return f"=COUNTA(Collection!{letter}:{letter})-1" if letter else 0

    def _counta_range(letter):
        return f"=COUNTA(Collection!{letter}2:{letter}{N})" if letter else 0

    completion = (
        f'=IFERROR(COUNTIF(Collection!{s}:{s},TRUE)/(COUNTA(Collection!{r}:{r})-1),0)'
        if s and r else 0
    )
    found_rate = (
        f'=IFERROR(COUNTIF(Collection!{f}:{f},TRUE)/COUNTIF(Collection!{s}:{s},TRUE),0)'
        if f and s else 0
    )

    TYPES = ["Full-length", "EP", "Single", "Live album", "Compilation", "Demo"]

    # Each entry: (label, value, bold_label, pct_format)
    rows_spec: list[tuple] = [
        ("Metal Archives Collection Statistics", None, True, False),
        (None, None, False, False),
        ("Collection", None, True, False),
        ("Total Releases",   _counta_col(r),           False, False),
        ("Unique Artists",   _unique_count(a) if a else 0, False, False),
        ("Searched",         _countif_col(s, "TRUE"),   False, False),
        ("Found",            _countif_col(f, "TRUE"),   False, False),
        ("Completion",       completion,                False, True),
        ("Found Rate",       found_rate,                False, True),
        (None, None, False, False),
        ("Release Types", None, True, False),
        *[(tp, f'=COUNTIF(Collection!{t}:{t},"{tp}")' if t else 0, False, False)
          for tp in TYPES],
        (None, None, False, False),
        ("Metadata", None, True, False),
        ("With Year",      _counta_range(y), False, False),
        ("With Genre",     _counta_range(g), False, False),
        ("Unique Genres",  _unique_count(g) if g else 0, False, False),
        (None, None, False, False),
        ("Other Sheets", None, True, False),
        ("Artists Processed",
            "=COUNTA(Artists!A:A)-1" if "Artists" in wb.sheetnames else 0,
            False, False),
        ("Requiring Review",
            "=COUNTA(Review!A:A)-1" if "Review" in wb.sheetnames else 0,
            False, False),
        ("Not Found",
            "=COUNTA('Not Found'!A:A)-1" if "Not Found" in wb.sheetnames else 0,
            False, False),
        (None, None, False, False),
        ("Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), False, False),
    ]

    if STATS_SHEET in wb.sheetnames:
        del wb[STATS_SHEET]
    ws = wb.create_sheet(STATS_SHEET)

    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=13)

    for row_num, (label, value, is_bold, is_pct) in enumerate(rows_spec, start=1):
        if label is None:
            ws.append([])
            continue
        ws.append([label] if value is None else [label, value])
        if is_bold:
            ws.cell(row=row_num, column=1).font = bold_font
        if is_pct and value:
            ws.cell(row=row_num, column=2).number_format = "0.0%"

    ws.cell(row=1, column=1).font = title_font
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16


def save_workbook(wb: Workbook, path: str):
    update_stats_sheet(wb)
    # Write to a temp file first, then rename into place atomically.
    # A same-filesystem rename on POSIX cannot be interrupted mid-write, so
    # KeyboardInterrupt or a crash during wb.save() leaves the original intact.
    tmp = str(Path(path).with_suffix(".tmp.xlsx"))
    try:
        wb.save(tmp)
        os.replace(tmp, path)
    except Exception as e:
        Path(tmp).unlink(missing_ok=True)
        backup = str(Path(path).with_suffix("")) + "_backup.xlsx"
        logger.error("Failed to save workbook to %s: %s. Trying backup: %s", path, e, backup)
        try:
            wb.save(backup)
            logger.info("Saved backup to %s", backup)
        except Exception as e2:
            logger.critical("Backup save also failed: %s", e2)
            raise
